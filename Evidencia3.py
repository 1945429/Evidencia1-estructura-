# Evidencia numero "3" Estructura de datos y su procesamineto

# Imports
import sys
import csv
import datetime
import os
import openpyxl
import sqlite3
from prettytable import PrettyTable
from sqlite3 import Error
# Limpieza de la terminal
os.system("cls")

# Creacion de la tabla "EJEMPLARES_REGISTRADOS"
try:
    with sqlite3.connect("EVIDENCIA_3.db") as conn:
        mi_cursor = conn.cursor()
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS EJEMPLARES_REGISTRADOS (ID_EJEMPLAR INTEGER PRIMARY KEY, TITULO TEXT NOT NULL, AUTOR TEXT NOT NULL, GENERO TEXT NOT NULL, AÑO_DE_PUBLICACION TEXT NOT NULL, ISBN TEXT NOT NULL, FECHA_DE_ADQUISICION TEXT NOT NULL);")
except Error as e:
    print(e)
except:
    print(f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
    regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
    os.system("cls")
finally:
    conn.close()

# Creacion de la tabla "AUTORES_REGISTRADOS"
try:
    with sqlite3.connect("EVIDENCIA_3.db") as conn:
        mi_cursor = conn.cursor()
        mi_cursor.execute(
            "CREATE TABLE IF NOT EXISTS AUTORES_REGISTRADOS (ID_AUTOR INTEGER PRIMARY KEY, APELLIDOS TEXT NOT NULL, NOMBRES TEXT NOT NULL);")
except Error as e:
    print(e)
except:
    print(f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
    regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
    os.system("cls")
finally:
    conn.close()

# Creacion de la tabla "GENEROS_REGISTRADOS"
try:
    with sqlite3.connect("EVIDENCIA_3.db") as conn:
        mi_cursor = conn.cursor()
        mi_cursor.execute(
            "CREATE TABLE IF NOT EXISTS GENEROS_REGISTRADOS (ID_GENERO INTEGER PRIMARY KEY, GENERO TEXT NOT NULL);")
except Error as e:
    print(e)
except:
    print(f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
    regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
    os.system("cls")
finally:
    conn.close()

# Verificacion de registros anteriores en la tabla "EJEMPLARES_REGISTRADOS"
try:
    with sqlite3.connect("EVIDENCIA_3.db") as conn:
        mi_cursor = conn.cursor()
        mi_cursor.execute("SELECT * FROM EJEMPLARES_REGISTRADOS")
        datos_previos = mi_cursor.fetchall()
except Error as e:
    print(e)
except:
    print(f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
    regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
    os.system("cls")
finally:
    conn.close()

# Registros encontrados
if datos_previos:
    pass
# Registros no encontrados
else:
    print("\nNO SE HA ENCONTRADO UNA VERSION DE DATOS PREVIA DE EJEMPLARES REGISTRADOS")
    incio = input("\nPULSA ENTER PARA DESPLEGAR EL MENU PRINCIPAL : ")
    os.system("cls")

# Menu principal
while True:
    print("\n  --- MENU PRINCIPAL ---")
    print("\n[1] REGISTRAR NUEVO EJEMPLAR")
    print("\n[2] CONSULTAS Y REPORTES")
    print("\n[3] REGISTRAR AUTOR")
    print("\n[4] REGISTRAR GENERO")
    print("\n[5] SALIR")
    opcion_menu_principal = input(
        "\n\nINGRESA EL NUMERO DE LA OPCION QUE DESEAS REALIZAR : ")
    os.system("cls")
    # Opcion 1 Menu principal
    if opcion_menu_principal == "1":
        while True:
            print("\n-- REGISTRAR NUEVO EJEMPLAR --")
            titulo = input("\nINGRESA EL TITULO DEL EJEMPLAR : ").upper()
            os.system("cls")
            while len(titulo.strip()) < 1 or titulo.isspace():
                print("\nEL TITULO DEL EJEMPLAR NO PUEDE SER OMITIDO")
                regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                os.system("cls")
                print("\n-- REGISTRAR NUEVO EJEMPLAR --")
                titulo = input("\nINGRESA EL TITULO DEL EJEMPLAR : ").upper()
                os.system("cls")
            print("\n-- REGISTRAR NUEVO EJEMPLAR --")
            print(f"\nTITULO DEL EJEMPLAR : {titulo}")
            print("\n[1] REPETIR REGISTRO")
            print("\n[2] CONFIRMAR REGISTRO")
            opcion = input(
                "\n\nINGRESA EL NUMERO DE LA OPCION QUE DESEAS REALIZAR : ")
            os.system("cls")
            if opcion == "1":
                pass
            elif opcion == "2":
                break
            else:
                print("\nEL DATO QUE INGRESO NO ES CORRECTO")
                regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                os.system("cls")

        while True:
            print("\n-- REGISTRAR NUEVO EJEMPLAR --")
            autor = input("\nINGRESA EL AUTOR DEL EJEMPLAR : ").upper()
            os.system("cls")

            while len(autor.strip()) < 1 or autor.isspace():
                print("\nEL AUTOR DEL EJEMPLAR NO PUEDE SER OMITIDO")
                regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                os.system("cls")
                print("\n-- REGISTRAR NUEVO EJEMPLAR --")
                autor = input("\nINGRESA EL AUTOR DEL EJEMPLAR : ").upper()
                os.system("cls")
            print("\n-- REGISTRAR NUEVO EJEMPLAR --")
            print(f"\nAUTOR DEL EJEMPLAR : {autor}")
            print("\n[1] REPETIR REGISTRO")
            print("\n[2] CONFIRMAR REGISTRO")
            opcion = input(
                "\n\nINGRESA EL NUMERO DE LA OPCION QUE DESEAS REALIZAR : ")
            os.system("cls")
            if opcion == "1":
                pass
            elif opcion == "2":
                break
            else:
                print("\nEL DATO QUE INGRESO NO ES CORRECTO")
                regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                os.system("cls")

        while True:
            print("\n-- REGISTRAR NUEVO EJEMPLAR --")
            genero = input("\nINGRESA EL GENERO DEL EJEMPLAR : ").upper()
            os.system("cls")
            while len(genero.strip()) < 1 or genero.isspace():
                print("\nEL GENERO DEL EJEMPLAR NO PUEDE SER OMITIDO")
                regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEvO : ")
                os.system("cls")
                print("\n-- REGISTRAR NUEVO EJEMPLAR --")
                genero = input("\nINGRESA EL GENERO DEL EJEMPLAR : ").upper()
                os.system("cls")

            print("\n-- REGISTRAR NUEVO EJEMPLAR --")
            print(f"\nGENERO DEL EJEMPLAR : {genero}")
            print("\n[1] REPETIR REGISTRO")
            print("\n[2] CONFIRMAR REGISTRO")
            opcion = input(
                "\n\nINGRESA EL NUMERO DE LA OPCION QUE DESEAS REALIZAR : ")
            os.system("cls")
            if opcion == "1":
                pass
            elif opcion == "2":
                break
            else:
                print("\nEL DATO QUE INGRESO NO ES CORRECTO")
                regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                os.system("cls")

        while True:
            while True:
                print("\n-- REGISTRAR NUEVO EJEMPLAR --")
                año_de_publicacion = input(
                    "\nINGRESA EL AÑO DE PUBLICACION DEL EJEMPLAR : ")
                os.system("cls")
                if año_de_publicacion.isdigit() and len(año_de_publicacion) == 4:
                    año_de_publicacion = int(año_de_publicacion)
                    if año_de_publicacion > 999 and año_de_publicacion <= 2030:
                        break
                    else:
                        print(
                            "\nEL AÑO DE PUBLICACION DEL EJEMPLAR NO PUEDE SER OMITIDO O MAYOR A 2030")
                        regreso = input(
                            "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                        os.system("cls")
                else:
                    print(
                        "\nEL AÑO DE PUBLICACION DEL EJEMPLAR NO PUEDE SER OMITIDO O MAYOR A 2030")
                    regreso = input(
                        "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                    os.system("cls")
            print("\n-- REGISTRAR NUEVO EJEMPLAR --")
            print(f"\nAÑO DE PUBLICACION DEL EJEMPLAR : {año_de_publicacion}")
            print("\n[1] REPETIR REGISTRO")
            print("\n[2] CONFIRMAR REGISTRO")
            opcion = input(
                "\n\nINGRESA EL NUMERO DE LA OPCION QUE DESEAS REALIZAR : ")
            os.system("cls")
            if opcion == "1":
                pass
            elif opcion == "2":
                break
            else:
                print("\nEL DATO QUE INGRESO NO ES CORRECTO")
                regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                os.system("cls")

        while True:
            print("\n-- REGISTRAR NUEVO EJEMPLAR --")
            isbn = input("\nINGRESA EL ISBN DEL EJEMPLAR : ").upper()
            os.system("cls")
            while len(isbn.strip()) < 1 or isbn.isspace():
                print("\nEL ISBN DEL EJEMPLAR NO PUEDE SER OMITIDO")
                regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                os.system("cls")
                print("\n-- REGISTRAR NUEVO EJEMPLAR --")
                isbn = input("\nINGRESA EL ISBN DEL EJEMPLAR : ").upper()
                os.system("cls")
            print("\n-- REGISTRAR NUEVO EJEMPLAR --")
            print(f"\nISBN DEL EJEMPLAR : {isbn}")
            print("\n[1] REPETIR REGISTRO")
            print("\n[2] CONFIRMAR REGISTRO")
            opcion = input(
                "\n\nINGRESA EL NUMERO DE LA OPCION QUE DESEAS REALIZAR : ")
            os.system("cls")
            if opcion == "1":
                pass
            elif opcion == "2":
                break
            else:
                print("\nEL DATO QUE INGRESO NO ES CORRECTO")
                regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                os.system("cls")

        while True:
            print("\n-- REGISTRAR NUEVO EJEMPLAR --")
            fecha_de_adquisicion = input(
                "\nINGRESA LA FECHA DE ADQUISICION DEL EJEMPLAR (DIA/MES/AÑO) : ")
            os.system("cls")

            try:
                fecha_convertida_2 = datetime.datetime.strptime(
                    fecha_de_adquisicion, '%d/%m/%Y').date()
            except ValueError:
                print("\nLA FECHA QUE INGRESO NO ES CORRECTA")
                regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                os.system("cls")
            else:
                print("\n-- REGISTRAR NUEVO EJEMPLAR --")
                print(
                    f"\nFECHA DE ADQUISICION DEL EJEMPLAR : {fecha_de_adquisicion}")
                print("\n[1] REPETIR REGISTRO")
                print("\n[2] CONFIRMAR REGISTRO")
                opcion = input(
                    "\n\nINGRESA EL NUMERO DE LA OPCION QUE DESEAS REALIZAR : ")
                os.system("cls")
                if opcion == "1":
                    pass
                elif opcion == "2":
                    break
                else:
                    print("\nEL DATO QUE INGRESO NO ES CORRECTO")
                    regreso = input(
                        "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                    os.system("cls")

        # Datos insertados en la tabla "EJEMPLARES_REGISTRADOS"
        try:
            with sqlite3.connect("EVIDENCIA_3.db") as conn:
                mi_cursor = conn.cursor()
                datos = {"TITULO": titulo, "AUTOR": autor, "GENERO": genero, "AÑO_DE_PUBLICACION":
                         año_de_publicacion, "ISBN": isbn, "FECHA_DE_ADQUISICION": fecha_de_adquisicion}
                mi_cursor.execute(
                    "INSERT INTO EJEMPLARES_REGISTRADOS (TITULO, AUTOR, GENERO, AÑO_DE_PUBLICACION, ISBN, FECHA_DE_ADQUISICION) VALUES (:TITULO, :AUTOR,:GENERO, :AÑO_DE_PUBLICACION, :ISBN, :FECHA_DE_ADQUISICION)", datos)

                # Id del ejemplar
                id_registro = mi_cursor.lastrowid
        except Error as e:
            print(e)
        except:
            print(f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
            regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
            os.system("cls")
        finally:
            conn.close()
        print("\nREGISTRO TERMINADO")
        print(f"\nLA CLAVE DEL REGISTRO ES : {id_registro}")
        regreso = input("\nPULSA ENTER PARA REGRESAR AL MENU PRINCIPAL : ")
        os.system("cls")

    # Opcion 2 Menu principal
    elif opcion_menu_principal == "2":
        # Menu consultas y reportes
        while True:
            print("\n-- CONSULTAS Y REPORTES --")
            print("\n[1] CONSULTAS")
            print("\n[2] REPORTES")
            print("\n[3] REGRESAR AL MENU PRINCIPAL")
            opcion_menu_consultas_y_reportes = input(
                "\n\nINGRESA EL NUMERO DE LA OPCION QUE DESEAS REALIZAR : ")
            os.system("cls")

            # Opcion 1 Menu consultas y reportes
            if opcion_menu_consultas_y_reportes == "1":
                # Menu consulta de titulo
                while True:
                    print("\n-- CONSULTAS --")
                    print("\n[1] POR TITULO")
                    print("\n[2] POR ISBN")
                    print("\n[3] REGRESAR AL MENU CONSULTAS Y REPORTES")
                    opcion_menu_consulta_de_titulo = input(
                        "\n\nINGRESA EL NUMERO DE LA OPCION QUE DESEAS REALIZAR : ")
                    os.system("cls")

                    # Opcion 1 Menu consulta de titulo
                    if opcion_menu_consulta_de_titulo == "1":
                        try:
                            with sqlite3.connect("EVIDENCIA_3.db") as conn:
                                mi_cursor = conn.cursor()
                                mi_cursor.execute(
                                    "SELECT * FROM EJEMPLARES_REGISTRADOS")
                                titulos_disponibles = mi_cursor.fetchall()
                        except Error as e:
                            print(e)
                        except:
                            print(
                                f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
                            regreso = input(
                                "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                            os.system("cls")
                        finally:
                            conn.close()
                        if titulos_disponibles:
                            print("\n-- CONSULTA POR TITULO --")
                            titulo = input(
                                "\nINGRESA EL TITULO DEL LIBRO QUE BUSCAS : ").upper()
                            os.system("cls")

                            try:
                                with sqlite3.connect("EVIDENCIA_3.db") as conn:
                                    mi_cursor = conn.cursor()
                                    datos = {"TITULO": titulo}
                                    mi_cursor.execute(
                                        "SELECT * FROM EJEMPLARES_REGISTRADOS WHERE TITULO = :TITULO", datos)
                                    titulo_disponible = mi_cursor.fetchall()
                            except Error as e:
                                print(e)
                            except:
                                print(
                                    f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
                                regreso = input(
                                    "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                                os.system("cls")
                            finally:
                                conn.close()
                            if titulo_disponible:
                                print("\nTITULO ENCONTRADO\n")
                                tabla = PrettyTable()
                                tabla.field_names = [
                                    "ID EJEMPLAR", "TITULO", "AUTOR", "GENERO", "AÑO DE PUBLICACION", "ISBN", "FECHA DE ADQUISICION"]
                                for dato1, dato2, dato3, dato4, dato5, dato6, dato7 in titulo_disponible:
                                    tabla.add_row(
                                        [dato1, dato2, dato3, dato4, dato5, dato6, dato7])
                                print(tabla)
                                regreso = input(
                                    "\nPULSA ENTER PARA REGRESAR AL MENU DE CONSULTAS : ")
                                os.system("cls")
                            else:
                                print("\nTITULO NO ENCONTRADO")
                                regreso = input(
                                    "\nPULSA ENTER PARA REGRESAR AL MENU DE CONSULTAS : ")
                                os.system("cls")
                        else:
                            print(
                                "\nACTUALMENTE NO HAY EJEMPLARES REGRISTRADOS EN LA BASE DE DATOS")
                            regreso = input(
                                "\nPULSA ENTER PARA REGRESAR AL MENU DE CONSULTAS : ")
                            os.system("cls")

                    # Opcion 2 Menu consulta de titulo
                    elif opcion_menu_consulta_de_titulo == "2":
                        try:
                            with sqlite3.connect("EVIDENCIA_3.db") as conn:
                                mi_cursor = conn.cursor()
                                mi_cursor.execute(
                                    "SELECT * FROM EJEMPLARES_REGISTRADOS")
                                titulos_disponibles = mi_cursor.fetchall()
                        except Error as e:
                            print(e)
                        except:
                            print(
                                f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
                            regreso = input(
                                "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                            os.system("cls")
                        finally:
                            conn.close()

                        if titulos_disponibles:
                            print("\n-- CONSULTA POR ISBN --")
                            isbn = input(
                                "\nINGRESA EL ISBN DEL LIBRO QUE BUSCAS : ").upper()
                            os.system("cls")

                            try:
                                with sqlite3.connect("EVIDENCIA_3.db") as conn:
                                    mi_cursor = conn.cursor()
                                    datos = {"ISBN": isbn}
                                    mi_cursor.execute(
                                        "SELECT * FROM EJEMPLARES_REGISTRADOS WHERE ISBN = :ISBN", datos)
                                    titulo_disponible = mi_cursor.fetchall()
                            except Error as e:
                                print(e)
                            except:
                                print(
                                    f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
                                regreso = input(
                                    "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                                os.system("cls")
                            finally:
                                conn.close()

                            if titulo_disponible:
                                print("\nISBN ENCONTRADO\n")
                                tabla = PrettyTable()
                                tabla.field_names = [
                                    "ID EJEMPLAR", "TITULO", "AUTOR", "GENERO", "AÑO DE PUBLICACION", "ISBN", "FECHA DE ADQUISICION"]
                                for dato1, dato2, dato3, dato4, dato5, dato6, dato7 in titulo_disponible:
                                    tabla.add_row(
                                        [dato1, dato2, dato3, dato4, dato5, dato6, dato7])
                                print(tabla)
                                regreso = input(
                                    "\nPULSA ENTER PARA REGRESAR AL MENU DE CONSULTAS : ")
                                os.system("cls")

                            else:
                                print("\nISBN NO ENCONTRADO")
                                regreso = input(
                                    "\nPULSA ENTER PARA REGRESAR AL MENU DE CONSULTAS : ")
                                os.system("cls")
                        else:
                            print(
                                "\nACTUALMENTE NO HAY EJEMPLARES REGRISTRADOS EN LA BASE DE DATOS")
                            regreso = input(
                                "\nPULSA ENTER PARA REGRESAR AL MENU DE CONSULTAS : ")
                            os.system("cls")

                    # Opcion 3 Menu consulta de titulo
                    elif opcion_menu_consulta_de_titulo == "3":
                        break
                    # Dato incorrecto Menu consultas y reportes
                    else:
                        print("\nEL DATO QUE INGRESO NO ES CORRECTO")
                        regreso = input(
                            "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                        os.system("cls")
            # Opcion 2 Menu consultas y reportes

            elif opcion_menu_consultas_y_reportes == "2":
                # Menu reportes
                while True:
                    print("\n-- REPORTES --")
                    print("\n[1] CATALOGO COMPLETO")
                    print("\n[2] REPORTE POR AUTOR")
                    print("\n[3] REPORTE POR GENERO")
                    print("\n[4] REPORTE POR AÑO DE PUBLICACION")
                    print("\n[5] REGRESAR AL MENU DE CONSULTAS Y REPORTES")
                    opcion_menu_reportes = input(
                        "\n\nINGRESA EL NUMERO DE LA OPCION QUE DESEAS REALIZAR : ")
                    os.system("cls")

                    # Opcion 1 Menu reportes
                    if opcion_menu_reportes == "1":
                        try:
                            with sqlite3.connect("EVIDENCIA_3.db") as conn:
                                mi_cursor = conn.cursor()
                                mi_cursor.execute(
                                    "SELECT * FROM EJEMPLARES_REGISTRADOS")
                                reporte_completo = mi_cursor.fetchall()
                        except Error as e:
                            print(e)
                        except:
                            print(
                                f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
                            regreso = input(
                                "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                            os.system("cls")
                        finally:
                            conn.close()
                        if reporte_completo:
                            tabla = PrettyTable()
                            tabla.field_names = [
                                "ID EJEMPLAR", "TITULO", "AUTOR", "GENERO", "AÑO DE PUBLICACION", "ISBN", "FECHA DE ADQUISICION"]
                            for dato1, dato2, dato3, dato4, dato5, dato6, dato7 in reporte_completo:
                                tabla.add_row(
                                    [dato1, dato2, dato3, dato4, dato5, dato6, dato7])
                            print("-- CATALOGO COMPLETO --\n")
                            print(tabla)
                            continuar = input(
                                "\nPULSA ENTER PARA CONTINUAR : ")
                            os.system("cls")
                            # Menu reportes exportados

                            while True:
                                print("-- REPORTES EXPORTADOS --")
                                print("\n[1]EXPORTAR REPORTE EN FORMATO CSV")
                                print("\n[2]EXPORTAR REPORTE EN FORMATO EXCEL")
                                print("\n[3]REGRESAR AL MENU DE REPORTES")
                                opcion_exportar = input(
                                    "\n\nINGRESA EL NUMERO DE LA OPCION QUE DESEAS REALIZAR : ")
                                os.system("cls")

                                # Opcion 1 Menu reportes exportados
                                if opcion_exportar == "1":
                                    print("\n-- EXPORTAR EN FORMATO CSV --")
                                    with open('REPORTE_COMPLETO.csv', mode='w', newline='') as archivo:
                                        writer = csv.writer(archivo)
                                        writer.writerow(reporte_completo)
                                    archivo.close()
                                    print("\nCSV GENERADO CORRECTAMENTE")
                                    print("\nINGRESA AL DISCO LOCAL C")
                                    print(
                                        "\nINGRESA A LA CARPETA EN DONDE SE UBICA EL ARCHIVO")
                                    regreso = input(
                                        "\nPULSA ENTER PARA REGRESAR AL MENU DE REPORTES EXPORTADOS : ")
                                    os.system("cls")

                                # Opcion 2 Menu reportes exportados
                                elif opcion_exportar == "2":
                                    print("\n-- EXPORTAR EN FORMATO EXCEL --")
                                    libro = openpyxl.Workbook()
                                    hoja = libro["Sheet"]
                                    hoja.title = "REPORTE COMPLETO"
                                    hoja["B2"].value = "ID EJEMPLAR"
                                    hoja["C2"].value = "TITULO"
                                    hoja["D2"].value = "AUTOR"
                                    hoja["E2"].value = "GENERO"
                                    hoja["F2"].value = "AÑO DE PUBLICACION"
                                    hoja["G2"].value = "ISBN"
                                    hoja["H2"].value = "FECHA DE ADQUISICION"
                                    for dato1, dato2, dato3, dato4, dato5, dato6, dato7 in reporte_completo:
                                        hoja.cell(row=dato1 + 3,
                                                  column=2).value = dato1
                                        hoja.cell(row=dato1 + 3,
                                                  column=3).value = dato2
                                        hoja.cell(row=dato1 + 3,
                                                  column=4).value = dato3
                                        hoja.cell(row=dato1 + 3,
                                                  column=5).value = dato4
                                        hoja.cell(row=dato1 + 3,
                                                  column=6).value = dato5
                                        hoja.cell(row=dato1 + 3,
                                                  column=7).value = dato6
                                        hoja.cell(row=dato1 + 3,
                                                  column=8).value = dato7
                                    libro.save("REPORTE_COMPLETO.xlsx")
                                    print("\nEXCEL GENERADO CORRECTAMENTE")
                                    print("\nINGRESA AL DISCO LOCAL C")
                                    print(
                                        "\nINGRESA A LA CARPETA EN DONDE SE UBICA EL ARCHIVO")
                                    regreso = input(
                                        "\nPULSA ENTER PARA REGRESAR AL MENU DE REPORTES EXPORTADOS : ")
                                    os.system("cls")

                                # Opcion 3 Menu reportes exportados
                                elif opcion_exportar == "3":
                                    break

                                # Dato incorrecto Menu reportes exportados
                                else:
                                    print("\nEL DATO QUE INGRESO NO ES CORRECTO")
                                    regreso = input(
                                        "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                                    os.system("cls")

                        else:
                            print(
                                "\nACTUALMENTE NO HAY EJEMPLARES REGRISTRADOS EN LA BASE DE DATOS")
                            regreso = input(
                                "\nPULSA ENTER PARA REGRESAR AL MENU DE REPORTES : ")
                            os.system("cls")

                    # Opcion 2 Menu reportes

                    elif opcion_menu_reportes == "2":

                        try:
                            with sqlite3.connect("EVIDENCIA_3.db") as conn:
                                mi_cursor = conn.cursor()
                                mi_cursor.execute(
                                    "SELECT * FROM EJEMPLARES_REGISTRADOS")
                                reporte_por_autor = mi_cursor.fetchall()
                        except Error as e:
                            print(e)
                        except:
                            print(
                                f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
                            regreso = input(
                                "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                            os.system("cls")
                        finally:
                            conn.close()
                        if reporte_por_autor:
                            try:

                                with sqlite3.connect("EVIDENCIA_3.db") as conn:
                                    mi_cursor = conn.cursor()
                                    mi_cursor.execute(
                                        "SELECT AUTOR FROM EJEMPLARES_REGISTRADOS")
                                    autores_disponibles = mi_cursor.fetchall()
                            except Error as e:
                                print(e)
                            except:
                                print(
                                    f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
                                regreso = input(
                                    "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                                os.system("cls")
                            finally:
                                conn.close()
                            lista_de_autores = []

                            for dato in autores_disponibles:
                                lista_de_autores.append(dato)
                            lista_de_autores = list(set(lista_de_autores))
                            tabla = PrettyTable()
                            tabla.field_names = ["AUTORES"]
                            for dato in lista_de_autores:
                                tabla.add_row([dato[0]])

                            print("\n-- REPORTE POR AUTOR --")
                            print("\nAUTORES DISPONIBLES\n")
                            print(tabla)
                            autor = input(
                                "\nINGRESA EL NOMBRE DEL AUTOR DEL CUAL DESEAS CONSULTAR SUS EJEMPLARES : ").upper()
                            os.system("cls")

                            try:
                                with sqlite3.connect("EVIDENCIA_3.db") as conn:
                                    mi_cursor = conn.cursor()
                                    datos = {"AUTOR": autor}
                                    mi_cursor.execute(
                                        "SELECT * FROM EJEMPLARES_REGISTRADOS WHERE AUTOR = :AUTOR", datos)
                                    ejemplares_de_autores = mi_cursor.fetchall()

                            except Error as e:
                                print(e)
                            except:
                                print(
                                    f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
                                regreso = input(
                                    "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                                os.system("cls")
                            finally:
                                conn.close()

                            if ejemplares_de_autores:
                                tabla = PrettyTable()
                                tabla.field_names = [
                                    "ID EJEMPLAR", "TITULO", "AUTOR", "GENERO", "AÑO DE PUBLICACION", "ISBN", "FECHA DE ADQUISICION"]
                                for dato1, dato2, dato3, dato4, dato5, dato6, dato7 in ejemplares_de_autores:
                                    tabla.add_row(
                                        [dato1, dato2, dato3, dato4, dato5, dato6, dato7])
                                print("\nAUTOR ENCONTRADO")
                                print(f"\nEJEMPLARES DEL AUTOR : {autor}\n")
                                print(tabla)
                                continuar = input(
                                    "\nPULSA ENTER PARA CONTINUAR : ")
                                os.system("cls")

                                # Menu reportes exportados
                                while True:
                                    print("-- REPORTES EXPORTADOS --")
                                    print(
                                        "\n[1]EXPORTAR REPORTE EN FORMATO CSV")
                                    print(
                                        "\n[2]EXPORTAR REPORTE EN FORMATO EXCEL")
                                    print("\n[3]REGRESAR AL MENU DE REPORTES")
                                    opcion_exportar = input(
                                        "\n\nINGRESA EL NUMERO DE LA OPCION QUE DESEAS REALIZAR : ")
                                    os.system("cls")

                                    # Opcion 1 Menu reportes exportados
                                    if opcion_exportar == "1":
                                        print("\n-- EXPORTAR EN FORMATO CSV --")
                                        with open('REPORTE_POR_AUTOR.csv', mode='w', newline='') as archivo:
                                            writer = csv.writer(archivo)
                                            writer.writerow(
                                                ejemplares_de_autores)
                                        archivo.close()
                                        print("\nCSV GENERADO CORRECTAMENTE")
                                        print("\nINGRESA AL DISCO LOCAL C")
                                        print(
                                            "\nINGRESA A LA CARPETA EN DONDE SE UBICA EL ARCHIVO")
                                        regreso = input(
                                            "\nPULSA ENTER PARA REGRESAR AL MENU DE REPORTES EXPORTADOS : ")
                                        os.system("cls")

                                    # Opcion 2 Menu reportes exportados
                                    elif opcion_exportar == "2":
                                        print(
                                            "\n-- EXPORTAR EN FORMATO EXCEL --")
                                        libro = openpyxl.Workbook()
                                        hoja = libro["Sheet"]
                                        hoja.title = "REPORTE POR AUTOR"
                                        hoja["B2"].value = "ID EJEMPLAR"
                                        hoja["C2"].value = "TITULO"
                                        hoja["D2"].value = "AUTOR"
                                        hoja["E2"].value = "GENERO"
                                        hoja["F2"].value = "AÑO DE PUBLICACION"
                                        hoja["G2"].value = "ISBN"
                                        hoja["H2"].value = "FECHA DE ADQUISICION"
                                        for dato1, dato2, dato3, dato4, dato5, dato6, dato7 in ejemplares_de_autores:
                                            hoja.cell(row=dato1 + 3,
                                                      column=2).value = dato1
                                            hoja.cell(row=dato1 + 3,
                                                      column=3).value = dato2
                                            hoja.cell(row=dato1 + 3,
                                                      column=4).value = dato3
                                            hoja.cell(row=dato1 + 3,
                                                      column=5).value = dato4
                                            hoja.cell(row=dato1 + 3,
                                                      column=6).value = dato5
                                            hoja.cell(row=dato1 + 3,
                                                      column=7).value = dato6
                                            hoja.cell(row=dato1 + 3,
                                                      column=8).value = dato7
                                        libro.save("REPORTE_POR_AUTOR.xlsx")
                                        print("\nEXCEL GENERADO CORRECTAMENTE")
                                        print("\nINGRESA AL DISCO LOCAL C")
                                        print(
                                            "\nINGRESA A LA CARPETA EN DONDE SE UBICA EL ARCHIVO")
                                        regreso = input(
                                            "\nPULSA ENTER PARA REGRESAR AL MENU DE REPORTES EXPORTADOS : ")
                                        os.system("cls")

                                    # Opcion 3 Menu reportes exportados
                                    elif opcion_exportar == "3":
                                        break

                                    # Dato incorrecto Menu reportes exportados
                                    else:
                                        print(
                                            "\nEL DATO QUE INGRESO NO ES CORRECTO")
                                        regreso = input(
                                            "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                                        os.system("cls")
                            else:
                                print("\nAUTOR NO ENCONTRADO")
                                regreso = input(
                                    "\nPULSA ENTER PARA REGRESAR AL MENU DE REPORTES : ")
                                os.system("cls")
                        else:
                            print(
                                "\nACTUALMENTE NO HAY EJEMPLARES REGRISTRADOS EN LA BASE DE DATOS")
                            regreso = input(
                                "\nPULSA ENTER PARA REGRESAR AL MENU DE REPORTES : ")
                            os.system("cls")

                    # Opcion 3 Menu reportes
                    elif opcion_menu_reportes == "3":
                        try:
                            with sqlite3.connect("EVIDENCIA_3.db") as conn:
                                mi_cursor = conn.cursor()
                                mi_cursor.execute(
                                    "SELECT * FROM EJEMPLARES_REGISTRADOS")
                                reporte_por_genero = mi_cursor.fetchall()
                        except Error as e:
                            print(e)
                        except:
                            print(
                                f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
                            regreso = input(
                                "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                            os.system("cls")
                        finally:
                            conn.close()
                        if reporte_por_genero:
                            try:
                                with sqlite3.connect("EVIDENCIA_3.db") as conn:
                                    mi_cursor = conn.cursor()
                                    mi_cursor.execute(
                                        "SELECT GENERO FROM EJEMPLARES_REGISTRADOS")
                                    generos_disponibles = mi_cursor.fetchall()

                            except Error as e:
                                print(e)
                            except:
                                print(
                                    f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
                                regreso = input(
                                    "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                                os.system("cls")
                            finally:
                                conn.close()
                            lista_de_generos = []
                            for dato in generos_disponibles:
                                lista_de_generos.append(dato)
                            lista_de_generos = list(set(lista_de_generos))
                            tabla = PrettyTable()
                            tabla.field_names = ["GENEROS"]
                            for dato in lista_de_generos:
                                tabla.add_row([dato[0]])
                            print("\n-- REPORTE POR GENERO --")
                            print("\nGENEROS DISPONIBLES\n")
                            print(tabla)
                            genero = input(
                                "\nINGRESA EL NOMBRE DEL GENERO DEL CUAL DESEAS CONSULTAR SUS EJEMPLARES : ").upper()
                            os.system("cls")

                            try:
                                with sqlite3.connect("EVIDENCIA_3.db") as conn:
                                    mi_cursor = conn.cursor()
                                    datos = {"GENERO": genero}
                                    mi_cursor.execute(
                                        "SELECT * FROM EJEMPLARES_REGISTRADOS WHERE GENERO = :GENERO", datos)
                                    ejemplares_de_generos = mi_cursor.fetchall()
                            except Error as e:
                                print(e)
                            except:
                                print(
                                    f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
                                regreso = input(
                                    "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                                os.system("cls")
                            finally:
                                conn.close()
                            if ejemplares_de_generos:
                                tabla = PrettyTable()
                                tabla.field_names = [
                                    "ID EJEMPLAR", "TITULO", "AUTOR", "GENERO", "AÑO DE PUBLICACION", "ISBN", "FECHA DE ADQUISICION"]
                                for dato1, dato2, dato3, dato4, dato5, dato6, dato7 in ejemplares_de_generos:
                                    tabla.add_row(
                                        [dato1, dato2, dato3, dato4, dato5, dato6, dato7])
                                print("\nGENERO ENCONTRADO")
                                print(f"\nEJEMPLARES DEL GENERO : {genero}\n")
                                print(tabla)
                                continuar = input(
                                    "\nPULSA ENTER PARA CONTINUAR : ")
                                os.system("cls")

                                # Menu reportes exportados
                                while True:
                                    print("-- REPORTES EXPORTADOS --")
                                    print(
                                        "\n[1]EXPORTAR REPORTE EN FORMATO CSV")
                                    print(
                                        "\n[2]EXPORTAR REPORTE EN FORMATO EXCEL")
                                    print("\n[3]REGRESAR AL MENU DE REPORTES")
                                    opcion_exportar = input(
                                        "\n\nINGRESA EL NUMERO DE LA OPCION QUE DESEAS REALIZAR : ")
                                    os.system("cls")

                                    # Opcion 1 Menu reportes exportados
                                    if opcion_exportar == "1":
                                        print("\n-- EXPORTAR EN FORMATO CSV --")
                                        with open('REPORTE_POR_GENERO.csv', mode='w', newline='') as archivo:
                                            writer = csv.writer(archivo)
                                            writer.writerow(
                                                ejemplares_de_generos)
                                        archivo.close()
                                        print("\nCSV GENERADO CORRECTAMENTE")
                                        print("\nINGRESA AL DISCO LOCAL C")
                                        print(
                                            "\nINGRESA A LA CARPETA EN DONDE SE UBICA EL ARCHIVO")
                                        regreso = input(
                                            "\nPULSA ENTER PARA REGRESAR AL MENU DE REPORTES EXPORTADOS : ")
                                        os.system("cls")

                                    # Opcion 2 Menu reportes exportados
                                    elif opcion_exportar == "2":
                                        print(
                                            "\n-- EXPORTAR EN FORMATO EXCEL --")
                                        libro = openpyxl.Workbook()
                                        hoja = libro["Sheet"]
                                        hoja.title = "REPORTE POR GENERO"
                                        hoja["B2"].value = "ID EJEMPLAR"
                                        hoja["C2"].value = "TITULO"
                                        hoja["D2"].value = "AUTOR"
                                        hoja["E2"].value = "GENERO"
                                        hoja["F2"].value = "AÑO DE PUBLICACION"
                                        hoja["G2"].value = "ISBN"
                                        hoja["H2"].value = "FECHA DE ADQUISICION"
                                        for dato1, dato2, dato3, dato4, dato5, dato6, dato7 in ejemplares_de_generos:
                                            hoja.cell(row=dato1 + 3,
                                                      column=2).value = dato1
                                            hoja.cell(row=dato1 + 3,
                                                      column=3).value = dato2
                                            hoja.cell(row=dato1 + 3,
                                                      column=4).value = dato3
                                            hoja.cell(row=dato1 + 3,
                                                      column=5).value = dato4
                                            hoja.cell(row=dato1 + 3,
                                                      column=6).value = dato5
                                            hoja.cell(row=dato1 + 3,
                                                      column=7).value = dato6
                                            hoja.cell(row=dato1 + 3,
                                                      column=8).value = dato7
                                        libro.save("REPORTE_POR_GENERO.xlsx")
                                        print("\nEXCEL GENERADO CORRECTAMENTE")
                                        print("\nINGRESA AL DISCO LOCAL C")
                                        print(
                                            "\nINGRESA A LA CARPETA EN DONDE SE UBICA EL ARCHIVO")
                                        regreso = input(
                                            "\nPULSA ENTER PARA REGRESAR AL MENU DE REPORTES EXPORTADOS : ")
                                        os.system("cls")

                                    # Opcion 3 Menu reportes exportados
                                    elif opcion_exportar == "3":
                                        break

                                    # Dato incorrecto Menu reportes exportados
                                    else:
                                        print(
                                            "\nEL DATO QUE INGRESO NO ES CORRECTO")
                                        regreso = input(
                                            "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                                        os.system("cls")
                            else:
                                print("\nGENERO NO ENCONTRADO")
                                regreso = input(
                                    "\nPULSA ENTER PARA REGRESAR AL MENU DE REPORTES : ")
                                os.system("cls")
                        else:
                            print(
                                "\nACTUALMENTE NO HAY EJEMPLARES REGRISTRADOS EN LA BASE DE DATOS")
                            regreso = input(
                                "\nPULSA ENTER PARA REGRESAR AL MENU DE REPORTES : ")
                            os.system("cls")

                    # Opcion 4 Menu reportes
                    elif opcion_menu_reportes == "4":
                        try:
                            with sqlite3.connect("EVIDENCIA_3.db") as conn:
                                mi_cursor = conn.cursor()
                                mi_cursor.execute(
                                    "SELECT * FROM EJEMPLARES_REGISTRADOS")
                                reporte_por_año_de_publicacion = mi_cursor.fetchall()
                        except Error as e:
                            print(e)
                        except:
                            print(
                                f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
                            regreso = input(
                                "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                            os.system("cls")
                        finally:
                            conn.close()

                        if reporte_por_año_de_publicacion:
                            try:
                                with sqlite3.connect("EVIDENCIA_3.db") as conn:
                                    mi_cursor = conn.cursor()
                                    mi_cursor.execute(
                                        "SELECT AÑO_DE_PUBLICACION FROM EJEMPLARES_REGISTRADOS")
                                    años_disponibles = mi_cursor.fetchall()
                            except Error as e:
                                print(e)
                            except:
                                print(
                                    f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
                                regreso = input(
                                    "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                                os.system("cls")
                            finally:
                                conn.close()
                            lista_de_años = []
                            for dato in años_disponibles:
                                lista_de_años.append(dato)
                            lista_de_años = list(set(lista_de_años))
                            tabla = PrettyTable()
                            tabla.field_names = ["AÑOS DE PUBLICACION"]
                            for dato in lista_de_años:
                                tabla.add_row([dato[0]])
                            print("\n-- REPORTE POR AÑO DE PUBLICACION --")
                            print("\nAÑOS DE PUBLICACION DISPONIBLES\n")
                            print(tabla)
                            año_de_publicacion = input(
                                "\nINGRESA EL AÑO DE PUBLICACION DEL CUAL DESEAS CONSULTAR SUS EJEMPLARES : ").upper()
                            os.system("cls")

                            try:
                                with sqlite3.connect("EVIDENCIA_3.db") as conn:
                                    mi_cursor = conn.cursor()
                                    datos = {
                                        "AÑO_DE_PUBLICACION": año_de_publicacion}
                                    mi_cursor.execute(
                                        "SELECT * FROM EJEMPLARES_REGISTRADOS WHERE AÑO_DE_PUBLICACION  = :AÑO_DE_PUBLICACION ", datos)
                                    ejemplares_de_años = mi_cursor.fetchall()
                            except Error as e:
                                print(e)
                            except:
                                print(
                                    f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
                                regreso = input(
                                    "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                                os.system("cls")
                            finally:
                                conn.close()
                            if ejemplares_de_años:
                                tabla = PrettyTable()
                                tabla.field_names = [
                                    "ID EJEMPLAR", "TITULO", "AUTOR", "GENERO", "AÑO DE PUBLICACION", "ISBN", "FECHA DE ADQUISICION"]
                                for dato1, dato2, dato3, dato4, dato5, dato6, dato7 in ejemplares_de_años:
                                    tabla.add_row(
                                        [dato1, dato2, dato3, dato4, dato5, dato6, dato7])
                                print("\nAÑO DE PUBLICACION ENCONTRADO")
                                print(
                                    f"\nEJEMPLARES DEL AÑO : {año_de_publicacion}\n")
                                print(tabla)
                                continuar = input(
                                    "\nPULSA ENTER PARA CONTINUAR : ")
                                os.system("cls")

                                # Menu reportes exportados
                                while True:
                                    print("-- REPORTES EXPORTADOS --")
                                    print(
                                        "\n[1]EXPORTAR REPORTE EN FORMATO CSV")
                                    print(
                                        "\n[2]EXPORTAR REPORTE EN FORMATO EXCEL")
                                    print("\n[3]REGRESAR AL MENU DE REPORTES")
                                    opcion_exportar = input(
                                        "\n\nINGRESA EL NUMERO DE LA OPCION QUE DESEAS REALIZAR : ")
                                    os.system("cls")

                                    # Opcion 1 Menu reportes exportados
                                    if opcion_exportar == "1":
                                        print("\n-- EXPORTAR EN FORMATO CSV --")
                                        with open('REPORTE_POR_AÑO_PUBLICACION.csv', mode='w', newline='') as archivo:
                                            writer = csv.writer(archivo)
                                            writer.writerow(ejemplares_de_años)
                                        archivo.close()
                                        print("\nCSV GENERADO CORRECTAMENTE")
                                        print("\nINGRESA AL DISCO LOCAL C")
                                        print(
                                            "\nINGRESA A LA CARPETA EN DONDE SE UBICA EL ARCHIVO")
                                        regreso = input(
                                            "\nPULSA ENTER PARA REGRESAR AL MENU DE REPORTES EXPORTADOS : ")
                                        os.system("cls")

                                    # Opcion 2 Menu reportes exportados
                                    elif opcion_exportar == "2":
                                        print(
                                            "\n-- EXPORTAR EN FORMATO EXCEL --")
                                        libro = openpyxl.Workbook()
                                        hoja = libro["Sheet"]
                                        hoja.title = "REPORTE POR AÑO DE PUBLICACION"
                                        hoja["B2"].value = "ID EJEMPLAR"
                                        hoja["C2"].value = "TITULO"
                                        hoja["D2"].value = "AUTOR"
                                        hoja["E2"].value = "GENERO"
                                        hoja["F2"].value = "AÑO DE PUBLICACION"
                                        hoja["G2"].value = "ISBN"
                                        hoja["H2"].value = "FECHA DE ADQUISICION"
                                        for dato1, dato2, dato3, dato4, dato5, dato6, dato7 in ejemplares_de_años:
                                            hoja.cell(row=dato1 + 3,
                                                      column=2).value = dato1
                                            hoja.cell(row=dato1 + 3,
                                                      column=3).value = dato2
                                            hoja.cell(row=dato1 + 3,
                                                      column=4).value = dato3
                                            hoja.cell(row=dato1 + 3,
                                                      column=5).value = dato4
                                            hoja.cell(row=dato1 + 3,
                                                      column=6).value = dato5
                                            hoja.cell(row=dato1 + 3,
                                                      column=7).value = dato6
                                            hoja.cell(row=dato1 + 3,
                                                      column=8).value = dato7
                                        libro.save(
                                            "REPORTE_POR_AÑO_DE_PUBLICACION.xlsx")
                                        print("\nEXCEL GENERADO CORRECTAMENTE")
                                        print("\nINGRESA AL DISCO LOCAL C")
                                        print(
                                            "\nINGRESA A LA CARPETA EN DONDE SE UBICA EL ARCHIVO")
                                        regreso = input(
                                            "\nPULSA ENTER PARA REGRESAR AL MENU DE REPORTES EXPORTADOS : ")
                                        os.system("cls")
                                    # Opcion 3 Menu reportes exportados
                                    elif opcion_exportar == "3":
                                        break
                                    # Dato incorrecto Menu reportes exportados
                                    else:
                                        print(
                                            "\nEL DATO QUE INGRESO NO ES CORRECTO")
                                        regreso = input(
                                            "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                                        os.system("cls")
                            else:
                                print("\nAÑO DE PUBLICACION NO ENCONTRADO")
                                regreso = input(
                                    "\nPULSA ENTER PARA REGRESAR AL MENU DE REPORTES : ")
                                os.system("cls")
                        else:
                            print(
                                "\nACTUALMENTE NO HAY EJEMPLARES REGRISTRADOS EN LA BASE DE DATOS")
                            regreso = input(
                                "\nPULSA ENTER PARA REGRESAR AL MENU DE REPORTES : ")
                            os.system("cls")

                    # Opcion 5 Menu reportes
                    elif opcion_menu_reportes == "5":
                        break

                    # Dato incorrecto Menu reportes
                    else:
                        print("\nEL DATO QUE INGRESO NO ES CORRECTO")
                        regreso = input(
                            "\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                        os.system("cls")

            # Opcion 3 Menu consultas y reportes
            elif opcion_menu_consultas_y_reportes == "3":
                break
            # Dato incorrecto Menu consultas y reportes
            else:
                print("\nEL DATO QUE INGRESO NO ES CORRECTO")
                regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                os.system("cls")

    # Opcion 3 Menu principal
    elif opcion_menu_principal == "3":
        while True:
            print("\n-- REGISTRAR AUTOR --")
            apellidos = input("\nINGRESA LOS APELLIDOS DEL AUTOR : ").upper()
            os.system("cls")
            while len(apellidos.strip()) < 1 or apellidos.isspace():
                print("\nLOS APELLIDOS DEL AUTOR NO PUEDEN SER OMITIDOS")
                regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                os.system("cls")
                print("\n-- REGISTRAR AUTOR --")
                apellidos = input(
                    "\nINGRESA LOS APELLIDOS DEL AUTOR : ").upper()
                os.system("cls")
            print("\n-- REGISTRAR AUTOR --")
            print(f"\nNOMBRES DEL AUTOR : {apellidos}")
            print("\n[1] REPETIR REGISTRO")
            print("\n[2] CONFIRMAR REGISTRO")
            opcion = input(
                "\n\nINGRESA EL NUMERO DE LA OPCION QUE DESEAS REALIZAR : ")
            os.system("cls")
            if opcion == "1":
                pass
            elif opcion == "2":
                break
            else:
                print("\nEL DATO QUE INGRESO NO ES CORRECTO")
                regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                os.system("cls")

        while True:
            print("\n-- REGISTRAR AUTOR --")
            nombres = input("\nINGRESA LOS NOMBRES DEL AUTOR : ").upper()
            os.system("cls")
            while len(nombres.strip()) < 1 or nombres.isspace():
                print("\nLOS NOMBRES DEL AUTOR NO PUEDEN SER OMITIDOS")
                regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                os.system("cls")
                print("\n-- REGISTRAR AUTOR --")
                nombres = input("\nINGRESA LOS NOMBRES DEL AUTOR : ").upper()
                os.system("cls")
            print("\n-- REGISTRAR AUTOR --")
            print(f"\nNOMBRES DEL AUTOR : {nombres}")
            print("\n[1] REPETIR REGISTRO")
            print("\n[2] CONFIRMAR REGISTRO")
            opcion = input(
                "\n\nINGRESA EL NUMERO DE LA OPCION QUE DESEAS REALIZAR : ")
            os.system("cls")
            if opcion == "1":
                pass
            elif opcion == "2":
                break
            else:
                print("\nEL DATO QUE INGRESO NO ES CORRECTO")
                regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                os.system("cls")

        # Datos insertados en la tabla "AUTORES_REGISTRADOS"
        try:
            with sqlite3.connect("EVIDENCIA_3.db") as conn:
                mi_cursor = conn.cursor()
                datos = {"APELLIDOS": apellidos, "NOMBRES": nombres, }
                mi_cursor.execute(
                    "INSERT INTO AUTORES_REGISTRADOS (APELLIDOS, NOMBRES) VALUES (:APELLIDOS, :NOMBRES)", datos)
                # Id del autor
                id_registro = mi_cursor.lastrowid
        except Error as e:
            print(e)
        except:
            print(f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
            regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
            os.system("cls")
        finally:
            conn.close()
        print("\nREGISTRO TERMINADO")
        print(f"\nLA CLAVE DEL REGISTRO ES : {id_registro}")
        regreso = input("\nPULSA ENTER PARA REGRESAR AL MENU PRINCIPAL : ")
        os.system("cls")

    # Opcion 4 Menu principal
    elif opcion_menu_principal == "4":
        while True:
            print("\n-- REGISTRAR GENERO --")
            genero = input("\nINGRESA EL GENERO : ").upper()
            os.system("cls")
            while len(genero.strip()) < 1 or genero.isspace():
                print("\nEL GENERO NO PUEDE SER OMITIDO")
                regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                os.system("cls")
                print("\n-- REGISTRAR GENERO --")
                genero = input("\nINGRESA EL GENERO DEL AUTOR : ").upper()
                os.system("cls")
            print("\n-- REGISTRAR GENERO --")
            print(f"\nGENERO : {genero}")
            print("\n[1] REPETIR REGISTRO")
            print("\n[2] CONFIRMAR REGISTRO")
            opcion = input(
                "\n\nINGRESA EL NUMERO DE LA OPCION QUE DESEAS REALIZAR : ")
            os.system("cls")

            if opcion == "1":
                pass
            elif opcion == "2":
                break
            else:
                print("\nEL DATO QUE INGRESO NO ES CORRECTO")
                regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
                os.system("cls")

        # Datos insertados en la tabla "GENEROS_REGISTRADOS"
        try:
            with sqlite3.connect("EVIDENCIA_3.db") as conn:
                mi_cursor = conn.cursor()
                datos = {"GENERO": genero}
                mi_cursor.execute(
                    "INSERT INTO GENEROS_REGISTRADOS (GENERO) VALUES (:GENERO)", datos)

                # Id del genero
                id_registro = mi_cursor.lastrowid
        except Error as e:
            print(e)

        except:
            print(f"SE PRODUJO EL SIGUIENTE ERROR : {sys.exc_info()[0]}")
            regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
            os.system("cls")
        finally:
            conn.close()
        print("\nREGISTRO TERMINADO")
        print(f"\nLA CLAVE DEL REGISTRO ES : {id_registro}")
        regreso = input("\nPULSA ENTER PARA REGRESAR AL MENU PRINCIPAL : ")
        os.system("cls")

    # Opcion 5 Menu principal
    elif opcion_menu_principal == "5":
        print("\nPROGRAMA FINALIZADO")
        print("\n")
        break

    # Dato incorrecto Menu principal
    else:
        print("\nEL DATO QUE INGRESO NO ES CORRECTO")
        regreso = input("\nPULSA ENTER PARA INTENTARLO DE NUEVO : ")
        os.system("cls")
