import os
import csv
import openpyxl

dic_libros = {}

try:
    with open("libros.csv", "r", newline="") as archivo:
        lector = csv.reader(archivo)
        next(lector)
        for clave, titulo, autor, genero, año_de_publicacion, isbn, fecha_adquisicion in lector:
            dic_libros[int(clave)] = (titulo, autor, genero, int(
                año_de_publicacion), int(isbn), fecha_adquisicion)
    print("SE RECUPERÒ UNA VERSION DE DATOS PREVIA.")
except Exception:
    print("NO SE ENCONTRO INFORMACION ANTERIOR, SE CONSIDERARÁ A ESTA LA PRIMERA EJECUCION DEL DESARROLLO")

while True:
    print("** MENU PRINCIPAL **")
    print("[1] REGISTRAR NUEVO EJEMPLAR")
    print("[2] CONSULTAS Y REPORTES")
    print("[3] SALIR")
    op = int(input("\nINGRESE EL NUMERO DE OPERACION QUE SE DESEA REALIZAR : "))
    os.system("cls")

    # Opcion 1 registrar un nuevo ejemplar.
    if op == 1:
        while True:

            print("REGISTRAR UN NUEVO EJEMPLAR")
            titulo = input("INGRESE EL TITULO DEL EJEMPLAR : ").upper()
            autor = input("INGRESE EL NOMBRE DEL AUTOR : ").upper()
            genero = input("INGRESE EL GENERO DEL LIBRO : ").upper()
            año_de_publicacion = int(input("INGRESE EL AÑO DE PUBLICACION : "))
            isbn = int(input("INGRESE EL ISBN : "))
            fecha_adquisicion = input(
                "INGRESE LA FECHA DE ADQUISICION EN EL FORMATO (DD/MM/AAAA) : ")

            # Clave del registro
            if dic_libros:
                clave = max(dic_libros) + 1
            else:
                clave = 1

            os.system("cls")
            print("\nREGISTRO TERMINADO")
            # Se guardan los datos en el diccionario dic_libros.
            dic_libros[clave] = titulo, autor, genero, año_de_publicacion, isbn, fecha_adquisicion
            respuesta = input(
                "DESEA SEGUIR AGREGANDO EJEMPLARES [SI/NO]").upper()
            if respuesta == "NO":
                os.system("cls")
                break

            os.system("cls")

    # Opcion de consultas y reportes.
    elif op == 2:
        # Sub menu "1"
        while True:
            print("** MENU DE CONSULTAS Y REPORTES **")
            print("[1] CONSULTA DE TITULO")
            print("[2] REPORTES")
            print("[3] VOLVER AL MENU PRINCIPAL ")
            op2 = int(
                input("\nINGRESE EL NUMERO DE OPERACION QUE SE DESEA REALIZAR : "))
            os.system("cls")

            if op2 == 1:
                # Sub menu "2"
                while True:
                    print("\n** MENU DE CONSULTA DE TITULO **")
                    print("[1] POR TITULO")
                    print("[2] POR ISBN")
                    print("[3] VOLVER AL MENU DE CONSULTAS Y REPORTES")
                    op3 = int(
                        input("\nINGRESE EL NUMERO DE OPERACION QUE DESEA REALIZAR : "))
                    os.system("cls")
                    # Opcion "1"
                    if op3 == 1:
                        print("LISTADO DE TITULOS DISPONIBLES: ")

                        for clave, datos in dic_libros.items():
                            print("*"*15)
                            print(" ", datos[0], "    ")
                            print("*"*15)

                        titulo = input(
                            "\nINGRESE EL TITULO DEL LIBRO QUE BUSCAS : ").upper()

                        os.system("cls")
                        print("*"*97)
                        print(
                            "CLAVE     TITULO      AUTOR      GENERO      AÑO DE PUBLICACION      ISBN      AÑO DE ADQUISICIÓN")
                        print("*"*97)
                        print("*"*97)

                        for clave, datos in dic_libros.items():
                            if titulo in datos:
                                clave_del_titulo = clave
                                print(" ", clave_del_titulo, "    ", dic_libros[clave_del_titulo][0], "  ", dic_libros[clave_del_titulo][1], "    ", dic_libros[clave_del_titulo][
                                      2], "            ", dic_libros[clave_del_titulo][3], "          ", dic_libros[clave_del_titulo][4], "      ", dic_libros[clave_del_titulo][5])
                                print("*"*97)

                    # Opcion "2"
                    elif op3 == 2:
                        print("LISTADO DE IBSN DISPONIBLES: ")

                        for clave, datos in dic_libros.items():
                            print("*"*15)
                            print(" ", datos[4], "    ")
                            print("*"*15)

                        isbn = int(
                            input("INGRESE EL ISBN DEL LIBRO QUE BUSCAS : "))
                        for clave, datos in dic_libros.items():
                            if isbn in datos:
                                clave_del_titulo = clave
                                print("*"*97)
                                print(
                                    "CLAVE     TITULO      AUTOR      GENERO      AÑO DE PUBLICACION      ISBN      AÑO DE ADQUISICIÓN")
                                print("*"*97)
                                print("*"*97)
                                print(" ", clave_del_titulo, "    ", dic_libros[clave_del_titulo][0], "  ", dic_libros[clave_del_titulo][1], "    ", dic_libros[clave_del_titulo][
                                      2], "            ", dic_libros[clave_del_titulo][3], "          ", dic_libros[clave_del_titulo][4], "      ", dic_libros[clave_del_titulo][5])
                                print("*"*97)

                    # Opcion "3"
                    elif op3 == 3:
                        os.system("cls")
                        break
                    # Error
                    else:
                        print("\nEL DATO QUE INGRESO ES INCORRECTO")
            elif op2 == 2:
                while True:
                    print("** MENU DE REPORTES **")
                    print("[1] CATALAGO COMPLETO")
                    print("[2] REPORTE POR AUTOR")
                    print("[3] REPORTE POR GENERO")
                    print("[4] REPORTE POR AÑO DE PUBLICACION")
                    print("[5] VOLVER AL MENU DE CONSULTAS Y REPORTES")
                    op4 = int(
                        input("\nINGRESE EL NUMERO DE OPERACION QUE DESEA REALIZAR : "))
                    os.system("cls")
                    if op4 == 1:
                        while True:
                            print("[1] EXPORTAR REPORTE EN FORMATO CSV.")
                            print("[2] EXPORTAR REPORTE EN FORMATO MsExcel.")
                            print("[3] NO EXPORTAR REPORTE")
                            print("[4] SALIR.")
                            opr = int(
                                input("QUE OPCION DE REPORTE DESEA REALIZAR ? "))

                            if opr == 1:
                                # Guarda el catalago de los libros  un archivo csv.
                                with open("catalagoLibros.csv", "w", newline="") as archivo:
                                    grabador = csv.writer(archivo)
                                    # Encabezados
                                    grabador.writerow(
                                        ("CLAVE", "TITULO", "AUTOR", "GENERO", "AÑO_PUBLICACION", "ISBN", "FECHA_ADQUISICION"))
                                    # Escribe el diccionario en el archivo csv.
                                    grabador.writerows(
                                        [(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5]) for clave, datos in dic_libros.items()])
                                print(
                                    "SE EXPORTO EL CATALAGO DE LIBROS EN UN ARCHIVO CSV.")

                            elif opr == 2:
                                # Reporte en Excel del catalago completo.
                                libro = openpyxl.Workbook()
                                hoja = libro["Sheet"]
                                hoja.title = "CatalagoLibros"
                                hoja.append(('CLAVE', 'TITULO', 'AUTOR', 'GENERO',
                                            'AÑO_PUBLICACION', 'ISBN', 'FECHA_ADQUISICION'))
                                for clave, datos in dic_libros.items():
                                    registro = (
                                        clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5])
                                    hoja.append(registro)

                                libro.save("CatalagoLibros.xlsx")
                                print(
                                    "SE EXPORTO EL CATALAGO DE LIBROS EN UN ARCHIVO DE EXCEL.")

                            elif opr == 3:
                                print("\nCATALAGO COMPLETO\n")
                                print("*"*97)
                                print(
                                    "CLAVE     TITULO      AUTOR      GENERO      AÑO DE PUBLICACION      ISBN      AÑO DE ADQUISICIÓN")
                                print("*"*97)
                                for clave, datos in dic_libros.items():
                                    print("*"*97)
                                    print(" ", clave, "    ", datos[0], "  ", datos[1], "    ", datos[2],
                                          "            ", datos[3], "          ", datos[4], "      ", datos[5])
                                    print("*"*97)
                            elif opr == 4:
                                break

                    elif op4 == 2:
                        # Reporte por autor
                        while True:
                            print("REPORTE POR AUTOR")
                            print("[1] EXPORTAR REPORTE EN FORMATO CSV.")
                            print("[2] EXPORTAR REPORTE EN FORMATO MsExcel.")
                            print("[3] NO EXPORTAR REPORTE")
                            print("[4] SALIR")
                            opr = int(
                                input("QUE OPCION DE REPORTE DESEA REALIZAR ? "))
                            os.system("cls")

                            if opr == 1:
                                print("LISTADO DE AUTORES DISPONIBLES: ")
                                for clave, datos in dic_libros.items():
                                    print("*"*15)
                                    print(" ", datos[1], "    ")
                                    print("*"*15)

                                autor = input(
                                    "INGRESA EL NOMBRE DEL AUTOR QUE BUSCAS : ").upper()

                                with open("reporte_autor.csv", "w", newline="") as archivo:
                                    grabador = csv.writer(archivo)
                                    # Encabezados
                                    grabador.writerow(
                                        ("CLAVE", "TITULO", "AUTOR", "GENERO", "AÑO_PUBLICACION", "ISBN", "FECHA_ADQUISICION"))
                                    # Escribe los libros del autor en el archivo csv.
                                    for clave, datos in list(dic_libros.items()):
                                        if datos[1] == autor:
                                            grabador.writerows(
                                                [(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5])])

                                print(
                                    f"SE EXPORTO UN REPORTE DE LIBROS DEL AUTOR {autor} EN UN ARCHIVO CSV.")
                            elif opr == 2:
                                print("LISTADO DE AUTORES DISPONIBLES: ")
                                for clave, datos in dic_libros.items():
                                    print("*"*15)
                                    print(" ", datos[1], "    ")
                                    print("*"*15)

                                autor = input(
                                    "INGRESA EL NOMBRE DEL AUTOR QUE BUSCAS : ").upper()
                                # Reporte en Excel de los libros del autor.
                                libro = openpyxl.Workbook()
                                hoja = libro["Sheet"]
                                hoja.title = "Libros"
                                hoja.append(('CLAVE', 'TITULO', 'AUTOR', 'GENERO',
                                            'AÑO_PUBLICACION', 'ISBN', 'FECHA_ADQUISICION'))
                                for clave, datos in list(dic_libros.items()):
                                    if datos[1] == autor:
                                        registro = (
                                            clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5])
                                        hoja.append(registro)

                                libro.save("LibrosAutor.xlsx")
                                print(
                                    f"SE EXPORTO EL REPORTE DE LIBROS DEL AUTOR {autor} EN UN ARCHIVO DE EXCEL.")

                            elif opr == 3:
                                print("LISTADO DE AUTORES DISPONIBLES: ")
                                for clave, datos in dic_libros.items():
                                    print("*"*15)
                                    print(" ", datos[1], "    ")
                                    print("*"*15)
                                autor = input(
                                    "INGRESA EL NOMBRE DEL AUTOR QUE BUSCAS : ").upper()
                                for clave, datos in list(dic_libros.items()):
                                    if datos[1] == autor:
                                        clave_del_titulo = clave
                                        print("*"*97)
                                        print(
                                            "CLAVE     TITULO      AUTOR      GENERO      AÑO DE PUBLICACION      ISBN      AÑO DE ADQUISICIÓN")
                                        print("*"*97)
                                        print("*"*97)
                                        print(" ", clave_del_titulo, "    ", dic_libros[clave_del_titulo][0], "  ", dic_libros[clave_del_titulo][1], "    ", dic_libros[clave_del_titulo][
                                              2], "            ", dic_libros[clave_del_titulo][3], "          ", dic_libros[clave_del_titulo][4], "      ", dic_libros[clave_del_titulo][5])
                                        print("*"*97)

                            elif opr == 4:
                                os.system("cls")
                                break
                    elif op4 == 3:
                        while True:
                            print("REPORTE POR GENERO")
                            print("[1] EXPORTAR REPORTE EN FORMATO CSV.")
                            print("[2] EXPORTAR REPORTE EN FORMATO MsExcel.")
                            print("[3] NO EXPORTAR REPORTE")
                            print("[4] SALIR")
                            opr = int(
                                input("QUE OPCION DE REPORTE DESEA REALIZAR ? "))
                            os.system("cls")
                            if opr == 1:
                                print("LISTADO DE GENEROS DISPONIBLES: ")
                                for clave, datos in dic_libros.items():
                                    print("*"*15)
                                    print(" ", datos[2], "    ")
                                    print("*"*15)

                                genero = input(
                                    "INGRESA EL NOMBRE DEL GENERO QUE BUSCAS: ").upper()

                                with open("reporte_genero.csv", "w", newline="") as archivo:
                                    grabador = csv.writer(archivo)
                                    # Encabezados
                                    grabador.writerow(
                                        ("CLAVE", "TITULO", "AUTOR", "GENERO", "AÑO_PUBLICACION", "ISBN", "FECHA_ADQUISICION"))
                                    # Escribe los libros del genero en el archivo csv.
                                    for clave, datos in list(dic_libros.items()):
                                        if datos[2] == genero:
                                            grabador.writerows(
                                                [(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5])])

                                print(
                                    f"SE EXPORTO UN REPORTE DE LIBROS DEL GENERO {genero} EN UN ARCHIVO CSV.")
                            elif opr == 2:
                                print("LISTADO DE GENEROS DISPONIBLES: ")
                                for clave, datos in dic_libros.items():
                                    print("*"*15)
                                    print(" ", datos[2], "    ")
                                    print("*"*15)

                                genero = input(
                                    "INGRESA EL NOMBRE DEL GENERO QUE BUSCAS: ").upper()

                                libro = openpyxl.Workbook()
                                hoja = libro["Sheet"]
                                hoja.title = "Libros"
                                hoja.append(('CLAVE', 'TITULO', 'AUTOR', 'GENERO',
                                            'AÑO_PUBLICACION', 'ISBN', 'FECHA_ADQUISICION'))
                                for clave, datos in list(dic_libros.items()):
                                    if datos[2] == genero:
                                        registro = (
                                            clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5])
                                        hoja.append(registro)

                                libro.save("LibrosGenero.xlsx")
                                print(
                                    f"SE EXPORTO EL REPORTE DE LIBROS DEL GENERO {genero} EN UN ARCHIVO DE EXCEL.")
                            elif opr == 3:
                                print("LISTADO DE GENEROS DISPONIBLES: ")
                                for clave, datos in dic_libros.items():
                                    print("*"*15)
                                    print(" ", datos[2], "    ")
                                    print("*"*15)
                                genero = input(
                                    "INGRESA EL NOMBRE DEL GENERO QUE BUSCAS: ").upper()
                                for clave, datos in list(dic_libros.items()):
                                    if datos[2] == genero:
                                        clave_del_titulo = clave
                                        print("*"*97)
                                        print(
                                            "CLAVE     TITULO      AUTOR      GENERO      AÑO DE PUBLICACION      ISBN      AÑO DE ADQUISICIÓN")
                                        print("*"*97)
                                        print("*"*97)
                                        print(" ", clave_del_titulo, "    ", dic_libros[clave_del_titulo][0], "  ", dic_libros[clave_del_titulo][1], "    ", dic_libros[clave_del_titulo][
                                              2], "            ", dic_libros[clave_del_titulo][3], "          ", dic_libros[clave_del_titulo][4], "      ", dic_libros[clave_del_titulo][5])
                                        print("*"*97)
                            elif opr == 4:
                                os.system("cls")
                                break
                    elif op4 == 4:
                        while True:
                            print("REPORTE POR AÑO DE PUBLICACION")
                            print("[1] EXPORTAR REPORTE EN FORMATO CSV.")
                            print("[2] EXPORTAR REPORTE EN FORMATO MsExcel.")
                            print("[3] NO EXPORTAR REPORTE")
                            print("[4] SALIR")
                            opr = int(
                                input("QUE OPCION DE REPORTE DESEA REALIZAR ? "))
                            os.system("cls")
                            if opr == 1:
                                print("LISTADO DE AÑOS DISPONIBLES: ")
                                for clave, datos in dic_libros.items():
                                    print("*"*15)
                                    print(" ", datos[3], "    ")
                                    print("*"*15)

                                año = int(
                                    input("INGRESA EL AÑO PARA CONSULTAR POR AÑO DE PUBLICACION: "))

                                with open("reporte_año.csv", "w", newline="") as archivo:
                                    grabador = csv.writer(archivo)
                                    # Encabezados
                                    grabador.writerow(
                                        ("CLAVE", "TITULO", "AUTOR", "GENERO", "AÑO_PUBLICACION", "ISBN", "FECHA_ADQUISICION"))
                                    # Escribe los libros del genero en el archivo csv.
                                    for clave, datos in list(dic_libros.items()):
                                        if int(datos[3]) == año:
                                            grabador.writerows(
                                                [(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5])])

                                print(
                                    f"SE EXPORTO UN REPORTE DE LIBROS DEL AÑO {año} EN UN ARCHIVO CSV.")
                            elif opr == 2:
                                for clave, datos in dic_libros.items():
                                    print("*"*15)
                                    print(" ", datos[3], "    ")
                                    print("*"*15)

                                año = int(
                                    input("INGRESA EL AÑO PARA CONSULTAR POR AÑO DE PUBLICACION: "))

                                libro = openpyxl.Workbook()
                                hoja = libro["Sheet"]
                                hoja.title = "Libros"
                                hoja.append(('CLAVE', 'TITULO', 'AUTOR', 'GENERO',
                                            'AÑO_PUBLICACION', 'ISBN', 'FECHA_ADQUISICION'))
                                for clave, datos in list(dic_libros.items()):
                                    if int(datos[3]) == año:
                                        registro = (
                                            clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5])
                                        hoja.append(registro)

                                libro.save("LibrosAño.xlsx")
                                print(
                                    f"SE EXPORTO EL REPORTE DE LIBROS DEL AÑO {año} EN UN ARCHIVO DE EXCEL.")
                            elif opr == 3:
                                for clave, datos in dic_libros.items():
                                    print("*"*15)
                                    print(" ", datos[3], "    ")
                                    print("*"*15)

                                año = int(
                                    input("INGRESA EL AÑO PARA CONSULTAR POR AÑO DE PUBLICACION: "))

                                for clave, datos in list(dic_libros.items()):
                                    if int(datos[3]) == año:
                                        clave_del_titulo = clave
                                        print("*"*97)
                                        print(
                                            "CLAVE     TITULO      AUTOR      GENERO      AÑO DE PUBLICACION      ISBN      AÑO DE ADQUISICIÓN")
                                        print("*"*97)
                                        print("*"*97)
                                        print(" ", clave_del_titulo, "    ", dic_libros[clave_del_titulo][0], "  ", dic_libros[clave_del_titulo][1], "    ", dic_libros[clave_del_titulo][
                                              2], "            ", dic_libros[clave_del_titulo][3], "          ", dic_libros[clave_del_titulo][4], "      ", dic_libros[clave_del_titulo][5])
                                        print("*"*97)

                            elif opr == 4:
                                os.system("cls")
                                break

                    elif op4 == 5:
                        os.system("cls")
                        break

            elif op2 == 3:
                os.system("cls")
                break
            # Error
            else:
                print("\nEL DATO QUE INGRESO ES INCORRECTO")

    # Opcion de salir
    elif op == 3:
        # Guarda los datos del diccionario dic_libros un archivo csv.
        with open("libros.csv", "w", newline="") as archivo:
            grabador = csv.writer(archivo)
            # Encabezados
            grabador.writerow(("CLAVE", "TITULO", "AUTOR", "GENERO",
                              "AÑO_PUBLICACION", "ISBN", "FECHA_ADQUISICION"))
            # Escribe el diccionario en el archivo csv.
            grabador.writerows([(clave, datos[0], datos[1], datos[2], datos[3],
                               datos[4], datos[5]) for clave, datos in dic_libros.items()])

        print("GRACIAS POR SU VISITA")
        print("\nPROGRAMA FINALIZADO\n")
        break
    else:
        print("EL DATO QUE INGRESO ES INCORRECTO")
